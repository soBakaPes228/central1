from fastapi import FastAPI, Request, Form, Query
from fastapi.responses import HTMLResponse, FileResponse, RedirectResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from datetime import datetime
import os
from database.init_db import get_connection, init_db

app = FastAPI(title="Табель учёта времени")

app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")


@app.on_event("startup")
def startup():
    init_db()


# ==================== HTML-СТРАНИЦЫ ====================

@app.get("/", response_class=HTMLResponse)
def employee_page(request: Request):
    conn = get_connection()
    users = conn.execute("SELECT * FROM users").fetchall()
    projects = conn.execute("SELECT * FROM projects").fetchall()
    conn.close()
    return templates.TemplateResponse("employee.html", {
        "request": request,
        "users": users,
        "projects": projects,
        "message": None
    })


@app.post("/employee/start", response_class=HTMLResponse)
def employee_start(request: Request, user_id: int = Form(...), project_id: int = Form(...)):
    conn = get_connection()
    now = datetime.now()
    today = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M:%S")

    open_record = conn.execute(
        "SELECT * FROM time_records WHERE user_id = ? AND end_time IS NULL AND date = ?",
        (user_id, today)
    ).fetchone()

    users = conn.execute("SELECT * FROM users").fetchall()
    projects = conn.execute("SELECT * FROM projects").fetchall()

    if open_record:
        conn.close()
        return templates.TemplateResponse("employee.html", {
            "request": request, "users": users, "projects": projects,
            "message": "❌ У вас уже есть незавершённая запись. Сначала завершите её."
        })

    conn.execute(
        "INSERT INTO time_records (user_id, project_id, date, start_time) VALUES (?, ?, ?, ?)",
        (user_id, project_id, today, time)
    )
    conn.commit()
    conn.close()

    return templates.TemplateResponse("employee.html", {
        "request": request, "users": users, "projects": projects,
        "message": f"✅ Начало работы зафиксировано: {time}"
    })


@app.post("/employee/stop", response_class=HTMLResponse)
def employee_stop(request: Request, user_id: int = Form(...)):
    conn = get_connection()
    now = datetime.now()
    today = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M:%S")

    open_record = conn.execute(
        "SELECT * FROM time_records WHERE user_id = ? AND end_time IS NULL AND date = ?",
        (user_id, today)
    ).fetchone()

    users = conn.execute("SELECT * FROM users").fetchall()
    projects = conn.execute("SELECT * FROM projects").fetchall()

    if not open_record:
        conn.close()
        return templates.TemplateResponse("employee.html", {
            "request": request, "users": users, "projects": projects,
            "message": "❌ Нечего завершать. Вы ещё не начинали работу."
        })

    conn.execute(
        "UPDATE time_records SET end_time = ? WHERE record_id = ?",
        (time, open_record["record_id"])
    )
    conn.commit()
    conn.close()

    return templates.TemplateResponse("employee.html", {
        "request": request, "users": users, "projects": projects,
        "message": f"✅ Работа завершена. Начало: {open_record['start_time']}, Конец: {time}"
    })


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(
    request: Request,
    project_id: int = Query(None),
    date_from: str = Query(None),
    date_to: str = Query(None)
):
    conn = get_connection()
    query = """
        SELECT r.record_id, u.full_name, p.name as project, r.date, r.start_time, r.end_time
        FROM time_records r
        JOIN users u ON r.user_id = u.user_id
        LEFT JOIN projects p ON r.project_id = p.project_id
        WHERE 1=1
    """
    params = []

    if project_id:
        query += " AND r.project_id = ?"
        params.append(project_id)
    if date_from:
        query += " AND r.date >= ?"
        params.append(date_from)
    if date_to:
        query += " AND r.date <= ?"
        params.append(date_to)

    query += " ORDER BY r.date DESC, r.start_time DESC"
    records = conn.execute(query, params).fetchall()
    projects = conn.execute("SELECT * FROM projects").fetchall()
    conn.close()

    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "records": records,
        "projects": projects,
        "selected_project": project_id,
        "date_from": date_from or "",
        "date_to": date_to or ""
    })


@app.get("/export")
def export_excel(project_id: int = Query(None), date_from: str = Query(None), date_to: str = Query(None)):
    from openpyxl import Workbook
    conn = get_connection()
    query = """
        SELECT u.full_name, p.name as project, r.date, r.start_time, r.end_time
        FROM time_records r
        JOIN users u ON r.user_id = u.user_id
        LEFT JOIN projects p ON r.project_id = p.project_id
        WHERE 1=1
    """
    params = []
    if project_id:
        query += " AND r.project_id = ?"
        params.append(project_id)
    if date_from:
        query += " AND r.date >= ?"
        params.append(date_from)
    if date_to:
        query += " AND r.date <= ?"
        params.append(date_to)

    query += " ORDER BY r.date DESC"
    records = conn.execute(query, params).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"
    for col, h in enumerate(["ФИО", "Проект", "Дата", "Начало", "Конец"], 1):
        ws.cell(row=1, column=col, value=h).font = __import__("openpyxl").styles.Font(bold=True)
    for row_idx, r in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=r["full_name"])
        ws.cell(row=row_idx, column=2, value=r["project"] or "—")
        ws.cell(row=row_idx, column=3, value=r["date"])
        ws.cell(row=row_idx, column=4, value=r["start_time"] or "—")
        ws.cell(row=row_idx, column=5, value=r["end_time"] or "не завершено")

    filename = f"otchet_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    filepath = os.path.join(os.path.dirname(__file__), filename)
    wb.save(filepath)
    return FileResponse(filepath, filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ==================== API (JSON) — CRUD ====================

# --- Пользователи ---
@app.get("/api/users")
def api_users():
    conn = get_connection()
    users = conn.execute("SELECT user_id, full_name FROM users").fetchall()
    conn.close()
    return [dict(u) for u in users]


@app.get("/api/users/{user_id}")
def api_user(user_id: int):
    conn = get_connection()
    user = conn.execute("SELECT * FROM users WHERE user_id = ?", (user_id,)).fetchone()
    conn.close()
    if not user:
        return JSONResponse({"error": "Пользователь не найден"}, 404)
    return dict(user)


@app.post("/api/users")
def api_user_create(full_name: str = Form(...)):
    conn = get_connection()
    cursor = conn.execute("INSERT INTO users (full_name) VALUES (?)", (full_name,))
    conn.commit()
    user_id = cursor.lastrowid
    conn.close()
    return {"user_id": user_id, "full_name": full_name}


@app.delete("/api/users/{user_id}")
def api_user_delete(user_id: int):
    conn = get_connection()
    conn.execute("DELETE FROM time_records WHERE user_id = ?", (user_id,))
    conn.execute("DELETE FROM users WHERE user_id = ?", (user_id,))
    conn.commit()
    conn.close()
    return {"status": "ok"}


# --- Проекты ---
@app.get("/api/projects")
def api_projects():
    conn = get_connection()
    projects = conn.execute("SELECT * FROM projects").fetchall()
    conn.close()
    return [dict(p) for p in projects]


@app.get("/api/projects/{project_id}")
def api_project(project_id: int):
    conn = get_connection()
    project = conn.execute("SELECT * FROM projects WHERE project_id = ?", (project_id,)).fetchone()
    conn.close()
    if not project:
        return JSONResponse({"error": "Проект не найден"}, 404)
    return dict(project)


@app.post("/api/projects")
def api_project_create(name: str = Form(...)):
    conn = get_connection()
    cursor = conn.execute("INSERT INTO projects (name) VALUES (?)", (name,))
    conn.commit()
    project_id = cursor.lastrowid
    conn.close()
    return {"project_id": project_id, "name": name}


@app.delete("/api/projects/{project_id}")
def api_project_delete(project_id: int):
    conn = get_connection()
    conn.execute("UPDATE time_records SET project_id = NULL WHERE project_id = ?", (project_id,))
    conn.execute("DELETE FROM projects WHERE project_id = ?", (project_id,))
    conn.commit()
    conn.close()
    return {"status": "ok"}


# --- Записи времени ---
@app.get("/api/records")
def api_records():
    conn = get_connection()
    records = conn.execute("""
        SELECT r.*, u.full_name, p.name as project_name
        FROM time_records r
        JOIN users u ON r.user_id = u.user_id
        LEFT JOIN projects p ON r.project_id = p.project_id
        ORDER BY r.date DESC
    """).fetchall()
    conn.close()
    return [dict(r) for r in records]


@app.get("/api/records/{record_id}")
def api_record(record_id: int):
    conn = get_connection()
    record = conn.execute("""
        SELECT r.*, u.full_name, p.name as project_name
        FROM time_records r
        JOIN users u ON r.user_id = u.user_id
        LEFT JOIN projects p ON r.project_id = p.project_id
        WHERE r.record_id = ?
    """, (record_id,)).fetchone()
    conn.close()
    if not record:
        return JSONResponse({"error": "Запись не найдена"}, 404)
    return dict(record)


@app.put("/api/records/{record_id}")
def api_record_update(record_id: int, project_id: int = Form(None), start_time: str = Form(None), end_time: str = Form(None)):
    conn = get_connection()
    if project_id is not None:
        conn.execute("UPDATE time_records SET project_id = ? WHERE record_id = ?", (project_id, record_id))
    if start_time is not None:
        conn.execute("UPDATE time_records SET start_time = ? WHERE record_id = ?", (start_time, record_id))
    if end_time is not None:
        conn.execute("UPDATE time_records SET end_time = ? WHERE record_id = ?", (end_time, record_id))
    conn.commit()
    conn.close()
    return {"status": "ok"}


@app.delete("/api/records/{record_id}")
def api_record_delete(record_id: int):
    conn = get_connection()
    conn.execute("DELETE FROM time_records WHERE record_id = ?", (record_id,))
    conn.commit()
    conn.close()
    return {"status": "ok"}
